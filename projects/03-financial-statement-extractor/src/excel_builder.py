"""
excel_builder.py
Builds a 4-sheet formatted Excel file from extracted financial statements.
Sheets: P&L, Balance Sheet, Cash Flow, Ratios
"""

from __future__ import annotations

import io

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# ── Style constants (matches Projects 1 & 2 dark theme) ──────────────────────

HEADER_FILL   = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
SECTION_FILL  = PatternFill(start_color="111827", end_color="111827", fill_type="solid")
ALT_ROW_FILL  = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")
RATIO_FILL    = PatternFill(start_color="EEF2FF", end_color="EEF2FF", fill_type="solid")

HEADER_FONT   = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
SECTION_FONT  = Font(name="Calibri", bold=True, color="818CF8", size=10)
LABEL_FONT    = Font(name="Calibri", size=10)
NUMBER_FONT   = Font(name="Calibri", size=10)
TITLE_FONT    = Font(name="Calibri", bold=True, size=13, color="1F2937")

THIN_BORDER = Border(
    bottom=Side(style="thin", color="E5E7EB"),
)

NUM_FORMAT    = '#,##0.00'
PCT_FORMAT    = '0.0%'
RATIO_FORMAT  = '0.00'
TEXT_FORMAT   = '@'


# ── Section rows (bold, no number) ───────────────────────────────────────────

_SECTION_LABELS = {
    "pnl": {"Gross Profit", "EBIT", "EBITDA", "Profit Before Tax", "Net Profit"},
    "bs":  {"Current Assets", "Non-Current Assets", "Total Assets",
            "Current Liabilities", "Non-Current Liabilities",
            "Total Liabilities", "Total Equity"},
    "cf":  {"Operating Cash Flow", "Investing Cash Flow",
            "Financing Cash Flow", "Net Change in Cash"},
}


def _is_section_row(label: str, stmt_type: str) -> bool:
    return label in _SECTION_LABELS.get(stmt_type, set())


# ── Sheet builder ─────────────────────────────────────────────────────────────

def _write_statement_sheet(
    ws,
    data: dict[str, dict[str, float | None]],
    stmt_type: str,
    company: str,
    sheet_title: str,
) -> None:
    years = []
    for item_data in data.values():
        for y in item_data:
            if y not in years:
                years.append(y)

    # Title row
    ws.cell(row=1, column=1).value = f"{company} — {sheet_title}"
    ws.cell(row=1, column=1).font = TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(years) + 1)

    # Sub-header: unit note
    ws.cell(row=2, column=1).value = "All figures in reporting currency as per source document"
    ws.cell(row=2, column=1).font = Font(name="Calibri", size=9, color="6B7280", italic=True)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(years) + 1)

    # Column headers
    ws.cell(row=3, column=1).value = "Line Item"
    ws.cell(row=3, column=1).fill = HEADER_FILL
    ws.cell(row=3, column=1).font = HEADER_FONT
    ws.cell(row=3, column=1).alignment = Alignment(horizontal="left", vertical="center")

    for col_idx, year in enumerate(years, start=2):
        cell = ws.cell(row=3, column=col_idx)
        cell.value = year
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="right", vertical="center")

    ws.row_dimensions[3].height = 22

    # Data rows
    for row_idx, (label, year_vals) in enumerate(data.items(), start=4):
        is_section = _is_section_row(label, stmt_type)
        is_other   = label.startswith("Other:")

        label_cell = ws.cell(row=row_idx, column=1)
        label_cell.value = label
        label_cell.font = SECTION_FONT if is_section else LABEL_FONT
        label_cell.alignment = Alignment(
            horizontal="left",
            indent=0 if is_section else (2 if not is_other else 4),
        )
        if is_section:
            label_cell.fill = SECTION_FILL
        label_cell.border = THIN_BORDER

        for col_idx, year in enumerate(years, start=2):
            val = year_vals.get(year)
            cell = ws.cell(row=row_idx, column=col_idx)
            if val is not None:
                cell.value = val
                cell.number_format = NUM_FORMAT
            else:
                cell.value = "—"
            cell.font = NUMBER_FONT
            cell.alignment = Alignment(horizontal="right")
            if is_section:
                cell.fill = SECTION_FILL
                cell.font = Font(name="Calibri", bold=True, size=10, color="818CF8")
            cell.border = THIN_BORDER

    # Column widths
    ws.column_dimensions["A"].width = 38
    for col_idx in range(2, len(years) + 2):
        ws.column_dimensions[get_column_letter(col_idx)].width = 16

    ws.freeze_panes = "B4"


# ── Ratios ────────────────────────────────────────────────────────────────────

def _safe_divide(num: float | None, den: float | None) -> float | None:
    if num is None or den is None or den == 0:
        return None
    return num / den


def _get_latest(data: dict[str, dict[str, float | None]], label: str) -> float | None:
    """Get the most recent year's value for a line item."""
    if label not in data:
        return None
    vals = list(data[label].values())
    for v in reversed(vals):
        if v is not None:
            return v
    return None


def _compute_ratios(
    pnl: dict,
    bs: dict,
    years: list[str],
) -> list[tuple[str, str, dict[str, str]]]:
    """
    Returns list of (category, ratio_name, {year: formatted_value}).
    """
    def get(data, label, year):
        return data.get(label, {}).get(year)

    def fmt_pct(v):
        if v is None:
            return "—"
        return f"{v*100:.1f}%"

    def fmt_x(v, decimals=2):
        if v is None:
            return "—"
        return f"{v:.{decimals}f}x"

    ratios = [
        # (category, name, numerator_stmt, numerator_label, denominator_stmt, denominator_label, format)
        ("Profitability", "Gross Margin",     "pnl", "Gross Profit",     "pnl", "Revenue",           "pct"),
        ("Profitability", "Operating Margin", "pnl", "EBIT",             "pnl", "Revenue",           "pct"),
        ("Profitability", "Net Margin",       "pnl", "Net Profit",       "pnl", "Revenue",           "pct"),
        ("Profitability", "ROE",              "pnl", "Net Profit",       "bs",  "Total Equity",      "pct"),
        ("Profitability", "ROA",              "pnl", "Net Profit",       "bs",  "Total Assets",      "pct"),
        ("Liquidity",     "Current Ratio",    "bs",  "Current Assets",   "bs",  "Current Liabilities","x"),
        ("Leverage",      "D/E Ratio",        "bs",  "Total Liabilities","bs",  "Total Equity",      "x"),
        ("Coverage",      "Interest Coverage","pnl", "EBIT",             "pnl", "Interest Expense",  "x"),
    ]

    stmts = {"pnl": pnl, "bs": bs}
    result = []
    for category, name, n_stmt, n_label, d_stmt, d_label, fmt in ratios:
        year_vals = {}
        for year in years:
            n = get(stmts[n_stmt], n_label, year)
            d = get(stmts[d_stmt], d_label, year)
            v = _safe_divide(n, d)
            year_vals[year] = fmt_pct(v) if fmt == "pct" else fmt_x(v)
        result.append((category, name, year_vals))

    return result


def _write_ratios_sheet(ws, pnl: dict, bs: dict, company: str) -> None:
    years = []
    for item_data in pnl.values():
        for y in item_data:
            if y not in years:
                years.append(y)

    ws.cell(row=1, column=1).value = f"{company} — Key Financial Ratios"
    ws.cell(row=1, column=1).font = TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(years) + 2)

    ws.cell(row=2, column=1).value = "Calculated from extracted P&L and Balance Sheet data"
    ws.cell(row=2, column=1).font = Font(name="Calibri", size=9, color="6B7280", italic=True)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(years) + 2)

    # Headers
    headers = ["Category", "Ratio"] + years
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col_idx)
        cell.value = h
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(
            horizontal="left" if col_idx <= 2 else "right",
            vertical="center",
        )
    ws.row_dimensions[3].height = 22

    ratios = _compute_ratios(pnl, bs, years)

    for row_idx, (category, name, year_vals) in enumerate(ratios, start=4):
        ws.cell(row=row_idx, column=1).value = category
        ws.cell(row=row_idx, column=1).font = Font(name="Calibri", size=10, color="6B7280")
        ws.cell(row=row_idx, column=2).value = name
        ws.cell(row=row_idx, column=2).font = Font(name="Calibri", bold=True, size=10)
        ws.cell(row=row_idx, column=2).fill = RATIO_FILL

        for col_idx, year in enumerate(years, start=3):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = year_vals.get(year, "—")
            cell.font = Font(name="Calibri", size=10)
            cell.alignment = Alignment(horizontal="right")
            cell.fill = RATIO_FILL
            cell.border = THIN_BORDER

        for col_idx in range(1, 3):
            ws.cell(row=row_idx, column=col_idx).border = THIN_BORDER

    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 24
    for col_idx in range(3, len(years) + 3):
        ws.column_dimensions[get_column_letter(col_idx)].width = 14

    ws.freeze_panes = "C4"


# ── Public API ────────────────────────────────────────────────────────────────

def build_excel(
    pnl: dict,
    bs: dict,
    cf: dict,
    company_name: str = "Company",
) -> bytes:
    """
    Build a 4-sheet Excel file and return as bytes.
    Each of pnl/bs/cf is a {line_item: {year: value}} dict.
    """
    output = io.BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        # Write placeholder DataFrames to create sheets (we'll overwrite with openpyxl)
        for sheet_name in ["P&L", "Balance Sheet", "Cash Flow", "Ratios"]:
            pd.DataFrame().to_excel(writer, sheet_name=sheet_name, index=False)

        wb = writer.book

        if pnl:
            _write_statement_sheet(wb["P&L"], pnl, "pnl", company_name, "Profit & Loss Statement")
        if bs:
            _write_statement_sheet(wb["Balance Sheet"], bs, "bs", company_name, "Balance Sheet")
        if cf:
            _write_statement_sheet(wb["Cash Flow"], cf, "cf", company_name, "Cash Flow Statement")
        if pnl and bs:
            _write_ratios_sheet(wb["Ratios"], pnl, bs, company_name)

    return output.getvalue()
